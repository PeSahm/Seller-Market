"""Render the bot fee report to an Excel (.xlsx) workbook.

This is the owner's primary deliverable: one row per successful bot BUY with
its matched/possible SELL and the realized fee, plus per-agent totals and a
raw-orders audit sheet. Money cells are real Excel numbers (not strings) so
the operator can sort/sum/pivot; dates are Excel dates.

Built with :mod:`openpyxl`. Returns raw ``bytes`` the route streams back.
"""

from __future__ import annotations

from datetime import datetime
from decimal import Decimal
from io import BytesIO
from typing import Optional
from uuid import UUID

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from app.models.broker_orders import BrokerOrder
from app.services.profit_report import FeeReport

_MONEY_FMT = "#,##0"
_DATE_FMT = "yyyy-mm-dd hh:mm:ss"
_PCT_FMT = "0.0000"
_HEADER_FILL = PatternFill("solid", fgColor="1F2937")
_HEADER_FONT = Font(bold=True, color="FFFFFF")

_SIDE_LABEL = {1: "Buy", 2: "Sell"}


def _num(value: Optional[Decimal]) -> Optional[float]:
    """Decimal → float for an Excel numeric cell. Rial magnitudes (~1e9–1e10)
    are well inside float64's exact-integer range, so no precision loss."""
    if value is None:
        return None
    return float(value)


def _xl_dt(dt: Optional[datetime]) -> Optional[datetime]:
    """openpyxl rejects tz-aware datetimes — drop tzinfo (the stored value is
    already the broker's wall clock)."""
    if dt is None:
        return None
    return dt.replace(tzinfo=None)


def _write_header(ws, headers: list[str]) -> None:
    ws.append(headers)
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
    ws.freeze_panes = "A2"


def _autosize(ws, widths: list[int]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def build_fee_workbook(
    report: FeeReport,
    orders: list[BrokerOrder],
    *,
    agent_names: dict[UUID, str],
    customer_names: dict[UUID, str],
    generated_for: str = "",
) -> bytes:
    """Render ``report`` (+ the raw ``orders``) into an .xlsx workbook."""
    wb = Workbook()

    # ---- Sheet 1: Buys & fees -------------------------------------------
    ws = wb.active
    ws.title = "Buys & fees"
    headers = [
        "Agent", "Customer", "Broker", "ISIN", "Symbol",
        "Buy date", "Buy volume", "Buy price", "Buy value (matched)",
        "Last sell date", "Sold volume", "Sell value",
        "Realized profit", "Fee %", "Fee amount", "Open volume", "Status",
    ]
    _write_header(ws, headers)
    for row in report.buy_rows:
        o = row.buy
        ws.append([
            agent_names.get(o.agent_id, "—") if o.agent_id else "—",
            customer_names.get(o.customer_id, o.account_username) if o.customer_id else o.account_username,
            o.broker,
            o.isin,
            o.symbol or o.symbol_title or "",
            _xl_dt(o.placed_at or o.created_at_broker),
            int(o.executed_volume or 0),
            _num(o.price),
            _num(row.buy_value),
            _xl_dt(row.last_sell_at),
            row.matched_volume,
            _num(row.sell_value),
            _num(row.realized_profit),
            _num(row.fee_percent),
            _num(row.fee),
            row.open_volume,
            row.status,
        ])
    _apply_formats(
        ws,
        date_cols=[6, 10],
        money_cols=[8, 9, 12, 13, 15],
        pct_cols=[14],
        nrows=len(report.buy_rows),
    )
    _autosize(ws, [16, 16, 10, 16, 10, 19, 12, 12, 18, 19, 12, 16, 16, 8, 16, 12, 10])

    # ---- Sheet 2: Per-agent totals --------------------------------------
    ws2 = wb.create_sheet("Per-agent totals")
    _write_header(
        ws2,
        ["Agent", "# buys", "Total buy value", "Realized profit",
         "Total fee", "Open volume"],
    )
    for totals in report.per_agent.values():
        ws2.append([
            agent_names.get(totals.agent_id, "—") if totals.agent_id else "—",
            totals.num_buys,
            _num(totals.total_buy_value),
            _num(totals.realized_profit),
            _num(totals.total_fee),
            totals.open_volume,
        ])
    # Grand total row.
    ws2.append([])
    ws2.append([
        "TOTAL", "", "", _num(report.grand_realized), _num(report.grand_fee), "",
    ])
    _apply_formats(ws2, date_cols=[], money_cols=[3, 4, 5], pct_cols=[], nrows=len(report.per_agent) + 2)
    _autosize(ws2, [18, 10, 18, 18, 18, 12])

    # ---- Sheet: Per-customer totals (owed − paid = remaining) (#116) -----
    wsc = wb.create_sheet("Per-customer totals")
    _write_header(
        wsc,
        ["Customer", "Agent", "# buys", "Realized profit",
         "Owed", "Paid", "Remaining", "Open volume"],
    )
    for t in report.per_customer.values():
        wsc.append([
            customer_names.get(t.customer_id, "—") if t.customer_id else "—",
            agent_names.get(t.agent_id, "—") if t.agent_id else "—",
            t.num_buys,
            _num(t.realized_profit),
            _num(t.total_fee),
            _num(t.paid),
            _num(t.remaining),
            t.open_volume,
        ])
    _apply_formats(wsc, date_cols=[], money_cols=[4, 5, 6, 7], pct_cols=[], nrows=len(report.per_customer))
    _autosize(wsc, [18, 16, 10, 18, 16, 16, 16, 12])

    # ---- Sheet: Realized remainder (mark-to-market, unsold > N days) ----
    if report.virtual_rows:
        wsv = wb.create_sheet("Realized remainder")
        _write_header(
            wsv,
            ["Customer", "Agent", "ISIN", "Symbol", "Open qty", "Oldest buy",
             "Avg buy price", "Price", "Trigger", "Status", "Fee"],
        )
        for v in report.virtual_rows:
            wsv.append([
                customer_names.get(v.customer_id, "—") if v.customer_id else "—",
                agent_names.get(v.agent_id, "—") if v.agent_id else "—",
                v.isin,
                v.symbol or "",
                int(v.open_qty),
                v.oldest_buy_date,  # real Excel date cell (None → empty)
                _num(v.avg_buy_price),
                int(v.price),
                "20d",
                "loss" if v.in_loss else "profit",
                _num(v.fee),
            ])
        _apply_formats(wsv, date_cols=[], money_cols=[7, 8, 11], pct_cols=[], nrows=len(report.virtual_rows))
        # Day-resolution format for Oldest buy (the shared _DATE_FMT carries a
        # time-of-day component that's noise for a pure date).
        for r in range(2, len(report.virtual_rows) + 2):
            wsv.cell(row=r, column=6).number_format = "yyyy-mm-dd"
        _autosize(wsv, [18, 16, 16, 12, 12, 12, 16, 14, 8, 8, 16])

    # ---- Sheet 3: Raw orders (audit) ------------------------------------
    ws3 = wb.create_sheet("Raw orders")
    _write_header(
        ws3,
        ["Tracking #", "Serial #", "Agent", "Customer", "Broker", "ISIN", "Symbol",
         "Side", "State", "Placed at", "Executed vol", "Price",
         "Executed amount", "Total fee", "Net traded value", "Bot?"],
    )
    for o in orders:
        ws3.append([
            o.tracking_number,
            getattr(o, "serial_number", None),
            agent_names.get(o.agent_id, "—") if o.agent_id else "—",
            customer_names.get(o.customer_id, o.account_username) if o.customer_id else o.account_username,
            o.broker,
            o.isin,
            o.symbol or "",
            _SIDE_LABEL.get(o.order_side, str(o.order_side)),
            o.state_desc or str(o.state),
            _xl_dt(o.placed_at or o.created_at_broker),
            int(o.executed_volume or 0),
            _num(o.price),
            _num(o.executed_amount),
            _num(o.total_fee),
            _num(o.net_traded_value),
            "yes" if o.is_bot else "",
        ])
    _apply_formats(ws3, date_cols=[10], money_cols=[12, 13, 14, 15], pct_cols=[], nrows=len(orders))
    _autosize(ws3, [14, 18, 16, 16, 10, 16, 10, 6, 18, 19, 12, 12, 16, 14, 16, 6])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _apply_formats(ws, *, date_cols, money_cols, pct_cols, nrows) -> None:
    """Apply number/date formats to data rows (row 2 .. nrows+1)."""
    for r in range(2, nrows + 2):
        for c in date_cols:
            ws.cell(row=r, column=c).number_format = _DATE_FMT
        for c in money_cols:
            ws.cell(row=r, column=c).number_format = _MONEY_FMT
        for c in pct_cols:
            ws.cell(row=r, column=c).number_format = _PCT_FMT


__all__ = ["build_fee_workbook"]
