"""Test the Excel fee report — the owner's primary deliverable.

Builds a small :class:`FeeReport` from in-memory ORM rows, renders the .xlsx,
then loads it back with openpyxl to assert the "Buys & fees" sheet has one row
per buy with real numeric money cells and the matched-sell + fee columns.
"""
from __future__ import annotations

import uuid
from datetime import date, datetime, timezone
from decimal import Decimal
from io import BytesIO

from openpyxl import load_workbook

from app.models.broker_orders import BrokerOrder
from app.services.fee_export import build_fee_workbook
from app.services.profit_report import (
    STATUS_REALIZED,
    AgentTotals,
    BuyFeeRow,
    CustomerFeeTotals,
    FeeReport,
    VirtualFeeRow,
)


def _buy_order(agent_id, customer_id):
    return BrokerOrder(
        customer_id=customer_id,
        agent_id=agent_id,
        broker="ayandeh",
        account_username="4580090306",
        tracking_number=909,
        isin="IRO1PNES0001",
        symbol="شپنا",
        order_side=1,
        price=Decimal("6310"),
        volume=200000,
        executed_volume=200000,
        executed_amount=Decimal("1266684544"),
        total_fee=Decimal("4684544"),
        net_traded_value=Decimal("1262000000"),
        state=3,
        state_desc="کاملا انجام شده",
        is_done=True,
        placed_at=datetime(2026, 6, 1, 8, 45, 1, tzinfo=timezone.utc),
        created_at_broker=datetime(2026, 6, 1, 8, 45, 0, tzinfo=timezone.utc),
        is_bot=True,
        raw_json={},
    )


def test_build_fee_workbook_buys_and_fees_sheet():
    agent_id = uuid.uuid4()
    customer_id = uuid.uuid4()
    buy = _buy_order(agent_id, customer_id)
    row = BuyFeeRow(
        buy=buy,
        matched_volume=200000,
        open_volume=0,
        buy_value=Decimal("1262000000"),
        sell_value=Decimal("1300000000"),
        realized_profit=Decimal("38000000"),
        fee=Decimal("380000"),
        fee_percent=Decimal("1.0"),
        last_sell_at=datetime(2026, 6, 2, 10, 0, 0, tzinfo=timezone.utc),
        sell_trackings=[1234],
        status=STATUS_REALIZED,
    )
    report = FeeReport(
        buy_rows=[row],
        per_agent={agent_id: AgentTotals(
            agent_id=agent_id, num_buys=1,
            total_buy_value=Decimal("1262000000"),
            realized_profit=Decimal("38000000"),
            total_fee=Decimal("380000"),
        )},
        per_customer={customer_id: CustomerFeeTotals(
            customer_id=customer_id, agent_id=agent_id, num_buys=1,
            total_buy_value=Decimal("1262000000"),
            realized_profit=Decimal("38000000"),
            total_fee=Decimal("380000"),
            paid=Decimal("80000"), remaining=Decimal("300000"),
        )},
        grand_realized=Decimal("38000000"),
        grand_fee=Decimal("380000"),
    )

    data = build_fee_workbook(
        report,
        [buy],
        agent_names={agent_id: "mostafa"},
        customer_names={customer_id: "Mostafa main"},
    )
    assert isinstance(data, bytes) and len(data) > 0

    wb = load_workbook(BytesIO(data))
    assert wb.sheetnames == [
        "Buys & fees", "Per-agent totals", "Per-customer totals", "Raw orders",
    ]
    cust = wb["Per-customer totals"]
    ch = [c.value for c in cust[1]]
    assert {"Owed", "Paid", "Remaining"} <= set(ch)
    crow = {ch[i]: cust[2][i].value for i in range(len(ch))}
    assert crow["Customer"] == "Mostafa main"
    assert crow["Owed"] == 380000 and crow["Paid"] == 80000 and crow["Remaining"] == 300000

    ws = wb["Buys & fees"]
    header = [c.value for c in ws[1]]
    assert "Realized profit" in header and "Fee amount" in header

    # One data row, real numeric money cells.
    data_row = {header[i]: ws[2][i].value for i in range(len(header))}
    assert data_row["Agent"] == "mostafa"
    assert data_row["Customer"] == "Mostafa main"
    assert data_row["ISIN"] == "IRO1PNES0001"
    assert data_row["Realized profit"] == 38000000
    assert isinstance(data_row["Fee amount"], (int, float))
    assert data_row["Fee amount"] == 380000
    assert data_row["Status"] == STATUS_REALIZED

    totals = wb["Per-agent totals"]
    tot_header = [c.value for c in totals[1]]
    assert "Total fee" in tot_header


def test_build_fee_workbook_empty_report_is_valid():
    data = build_fee_workbook(
        FeeReport(), [], agent_names={}, customer_names={}
    )
    wb = load_workbook(BytesIO(data))
    assert "Buys & fees" in wb.sheetnames
    # header only, no data rows
    assert wb["Buys & fees"].max_row == 1


def test_build_fee_workbook_mark_to_market_sheet():
    # A 20d virtual row renders with the Oldest-buy column and the "20d"
    # trigger label (the sell-trigger label is gone with the FIFO revert).
    agent_id = uuid.uuid4()
    customer_id = uuid.uuid4()
    report = FeeReport(
        virtual_rows=[VirtualFeeRow(
            customer_id=customer_id, agent_id=agent_id, broker="ayandeh",
            isin="IRO1PNES0001", symbol="شپنا", open_qty=40,
            avg_buy_price=Decimal("6000"), price=7000, trigger="20d",
            in_loss=False, fee=Decimal("400"),
            oldest_buy_date=date(2026, 6, 1),
        )],
    )
    data = build_fee_workbook(
        report, [],
        agent_names={agent_id: "mostafa"},
        customer_names={customer_id: "Mostafa main"},
    )
    wb = load_workbook(BytesIO(data))
    assert "Realized remainder" in wb.sheetnames
    ws = wb["Realized remainder"]
    header = [c.value for c in ws[1]]
    assert "Oldest buy" in header
    row = {header[i]: ws[2][i].value for i in range(len(header))}
    assert row["Customer"] == "Mostafa main"
    assert row["Trigger"] == "20d"
    assert row["Oldest buy"] == "2026-06-01"
    assert row["Open qty"] == 40 and row["Fee"] == 400
